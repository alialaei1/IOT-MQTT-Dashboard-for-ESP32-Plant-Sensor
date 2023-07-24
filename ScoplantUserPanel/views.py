import os
from django.http.response import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .forms import *
from .models import *
from ScoplantDevices.models import *
from ScoplantLogInfo.models import *
from random import randint
from jdatetime import date
from jalaali import Jalaali
from django.shortcuts import render
from django.http import HttpResponse
import pathlib
from django.http import FileResponse
from django.template.loader import get_template
# from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import PatternFill, Font
from .mqtt import subscriber, publisher
import plotly
import plotly.express as px
import pandas as pd
import plotly.graph_objs as go
import datetime


# 404 Custom Error
def error_404_view(request, exception):
    return render(request, '404.html')


# 500 Custom Error
def error_500_view(request):
    return render(request, '500.html')


subscriber()


@login_required(login_url='/login')
def add_device(request):
    # Date of today in Jalali
    now_date = date.today()
    user_id = request.user.id
    user = User.objects.get(id=user_id)
    add_new_device = AddNewDevice(request.POST or None)
    device_info = AddDeviceInfo.objects.all()

    # activated_before is a var to see if a Device has been activate before or not
    activated_before = False
    # wrong_information is a var to see if a Device info is worong or not
    wrong_information = False
    # invalid user is flag for error char and show in add-device
    invalid_username = False

    if request.method == "POST":
        Username = request.POST['Username']

        # REGEX for checking username for invalid char
        if "#" in Username or "$" in Username:
            invalid_username = True

        Version = request.POST['Version']
        Name = request.POST['Name']
        Location = request.POST['Location']
        # Checks if there is a Device with this Information
        if AccountDevice.objects.filter(Username=Username, Version=Version).first():
            # q1 is the device that you entred its Information
            q1 = AccountDevice.objects.filter(
                Username=Username, Version=Version).first()
            # Checks if q1 is activated Before or not
            if q1.Active == False:
                # if q1 is not activated before it wil add to users device list
                AddDeviceInfo.objects.create(Username=Username, Version=Version,
                                             Name=Name, Location=Location, id=randint(0, 9999999999), User=user, Date=now_date)
                # It will set active filed to True
                q1.Active = True
                q1.save(update_fields=['Active'])
                # It will save the changes
            # else q1 activated before
            else:
                activated_before = True
        # else there is such a device
        else:
            wrong_information = True
    # it will store number of all devices of a user
    number_of_devices = 0
    for item in device_info:
        if item.User == user:
            number_of_devices = number_of_devices+1
    avatar_gen = f"/site_statics/images/Avatars/Avatar {randint(1,12)}.jpg"

    context = {
        'add_device': add_new_device,
        'device_info': device_info,
        'user': user,
        'number_of_devices': number_of_devices,
        'title': "Device Registration",
        'invalid_username': invalid_username,
        'activated_before': activated_before,
        'wrong_information': wrong_information,
        'avatar_gen_url': avatar_gen,
    }
    return render(request, 'add-device.html', context)


@login_required(login_url='/login')
def device_info(request, device_id):
    device_info = AddDeviceInfo.objects.all()
    user_id = request.user.id
    user = User.objects.get(id=user_id)
    data = None
    try:
        data = LogInfo.objects.filter(id_device=device_id).latest("Time_Log")
    except:
        pass
    date_of_today = datetime.date.today()
    last10days = datetime.datetime(date_of_today.year, date_of_today.month,  date_of_today.day -9)
    labels =[]
    Lux_Log = []
    Humidity_Log = []
    Temperature_Log = []
    SoilMoisture_Log = []
    SoilTemperature_Log = []
    EC_Log = []

    LogsQS = LogInfo.objects.filter(id_device=device_id).filter(Date_Log__range=[last10days, date_of_today])

    for entry in LogsQS:
        labels.append(str(entry.Date_Log))

    for each_data in LogsQS:
        Lux_Log.append(each_data.Lux_Log)
        Humidity_Log.append(each_data.Humidity_Log)
        Temperature_Log.append(each_data.Temperature_Log)
        SoilMoisture_Log.append(each_data.SoilMoisture_Log)
        SoilTemperature_Log.append(each_data.SoilTemperature_Log)
        EC_Log.append(each_data.EC_Log)
    lstRecords = []
    for record in LogsQS:
        # Convert gregorian date into jalali date
        DateLOG = str(record.Date_Log)
        JDateLOG = Jalaali.to_jalaali(
            int(DateLOG[0:4]), int(DateLOG[5:7]), int(DateLOG[8:10]))
        JDateLOG = f"{JDateLOG['jy']}-{JDateLOG['jm']}-{JDateLOG['jd']}"
        each_data = f"{record.Date_Log}, {record.Time_Log}, {record.Lux_Log}, {record.Humidity_Log}, {record.Temperature_Log}, {record.SoilMoisture_Log}, {record.SoilTemperature_Log}, {record.EC_Log}"

        lstRecords.append(each_data)

    
    # checks if device_id in url is exists or not
    if device_id is not None:
        # deviceqs is a list of all device of the current user
        deviceqs = AddDeviceInfo.objects.get_queryset().get(
            id=device_id, User=request.user.id)
    # else no device_id entred will return 404 Not Found
    else:
        raise Http404()
    add_new_device = AddNewDevice(request.POST or None)

    # Checks if form request is POST
    if request.method == "POST":
        # Checks if add new device info is valid an entered currectly
        if add_new_device.is_valid:
            Sampling_Rate = request.POST['Sampling_Rate']
            deviceqs.Sampling_Rate = Sampling_Rate
            deviceqs.save()
            publisher(Sampling_Rate)
    # it create a right to left jalali date
    rtl_date = deviceqs.Date
    date_part = rtl_date.split("-")
    final_date = date_part[2] + "-" + date_part[1] + "-" + date_part[0]

    number_of_devices = 0
    for item in device_info:
        if item.User == user:
            number_of_devices = number_of_devices+1

    avatar_gen = f"/site_statics/images/Avatars/Avatar {randint(1,12)}.jpg"

    context = {
        'device_info': device_info,
        'log_sample': add_new_device,
        'number_of_devices': number_of_devices,
        'title': "Measuring Information",
        'user': user,
        'device_id': device_id,
        'deviceqs': deviceqs,
        'date_fa': final_date,
        'avatar_gen_url': avatar_gen,
        'last_data': data,
        'labels': labels,
        'Lux_Log':Lux_Log,
        'Humidity_Log':Humidity_Log,
        'Temperature_Log':Temperature_Log,
        'SoilMoisture_Log':SoilMoisture_Log,
        'SoilTemperature_Log':SoilTemperature_Log,
        'EC_Log':EC_Log,
    }
    return render(request, 'info-device.html', context)


@login_required(login_url='/login')
def remove_device(request, device_id):
    # checks if device_id in url is exists or not
    if device_id is not None:
        # deviceqs is the device that in url entered
        deviceqs = AddDeviceInfo.objects.get_queryset().get(
            id=device_id, User=request.user.id)
        if deviceqs is not None:
            # q1 is the account of the device that selected to be removed
            q1 = AccountDevice.objects.filter(
                Username=deviceqs.Username).first()
            # remove the selected device
            deviceqs.delete()
            # Deactive the seleced device
            q1.Active = False
            # It will save the changes
            q1.save()
        # After remoing the device it will redirect to home page
        return redirect('/')
    # if device id was not entred it will raise 404 Error
    raise Http404()


@login_required(login_url='/login')
def reporting_device(request, device_id):
    user_id = request.user.id
    user = User.objects.get(id=user_id)
    if AddDeviceInfo.objects.get_queryset().get(id=device_id, User=request.user.id):
        pass
    else:
        return Http404()
    device_info = AddDeviceInfo.objects.all()
    export_methods = ExportingMethods(request.POST or None)
    # wrong_date_range if False Date Range is currect if True Data Range is Wrong
    wrong_date_range = False

    # Checks if form request is POST
    if request.method == "POST":
        # startDate is the selected start date in form
        startDate = request.POST.get("start_date").split("/")
        # gregorianStartDate is startDate Converted to gregorian
        gregorianStartDate = date(int(startDate[0]), int(
            startDate[1]), int(startDate[2])).togregorian()

        # endDate is the selected end date in form
        endDate = request.POST.get("end_date").split("/")
        # gregorianEndDate is endDate Converted to gregorian
        gregorianEndtDate = date(int(endDate[0]), int(
            endDate[1]), int(endDate[2])).togregorian()
        # selectExport is the dropdown that select PDF, Excel, ...
        selectExport = request.POST.get("Select_Export")
        # selectParameter is the dropdown that select Lux, Ec, ...
        selectParameter = request.POST.get("Select_Parameter")

        # Checks if Selected Start Sate is smaller than Selected End Date
        if not(gregorianStartDate > gregorianEndtDate):
            # checks if device_id in url is exists or not
            if device_id is not None:
                # selected_device is the device that in url entered
                selected_device = AddDeviceInfo.objects.get_queryset().get(
                    id=device_id, User=request.user.id)
                # Checks if selected_device is exists or not
                if selected_device is not None:
                    # LogQS is the List of Logs for the selcted device and range
                    LogsQS = LogInfo.objects.filter(id_device=selected_device).filter(
                        Date_Log__range=[gregorianStartDate, gregorianEndtDate])
            # Checks if selectExport is ExcelTableExport
            if selectExport == "ExcelTableExport":
                # lstRecords it will be the list of all Queris that exists in LogQS
                lstRecords = []
                # This Loop is appending LogQS values to lstRecords
                for record in LogsQS:
                    # Convert gregorian date into jalali date
                    DateLOG = str(record.Date_Log)
                    JDateLOG = Jalaali.to_jalaali(
                        int(DateLOG[0:4]), int(DateLOG[5:7]), int(DateLOG[8:10]))
                    JDateLOG = f"{JDateLOG['jy']}-{JDateLOG['jm']}-{JDateLOG['jd']}"

                    # Remove ms from Time_Log
                    TimeLOG = record.Time_Log

                    # This if elif conditions checks the Selcted Parameter
                    if selectParameter == "Lux":
                        Parameter = record.Lux_Log

                    elif selectParameter == "Humidity":
                        Parameter = record.Humidity_Log

                    elif selectParameter == "Temperature":
                        Parameter = record.Temperature_Log

                    elif selectParameter == "Soil_Moisture":
                        Parameter = record.SoilMoisture_Log

                    elif selectParameter == "Soil_tempurature":
                        Parameter = record.SoilTemperature_Log

                    elif selectParameter == "EC":
                        Parameter = record.EC_Log

                    elif selectParameter == "Total":
                        each_list = []
                        each_list.append(JDateLOG)
                        each_list.append(record.Time_Log)
                        each_list.append(record.Lux_Log)
                        each_list.append(record.Humidity_Log)
                        each_list.append(record.Temperature_Log)
                        each_list.append(record.SoilMoisture_Log)
                        each_list.append(record.SoilTemperature_Log)
                        each_list.append(record.EC_Log)
                        # append each query record as a list to lstRecords
                        lstRecords.append(each_list)

                    # each_data is a list of a query record
                    if selectParameter == "Total":
                        pass
                    else:
                        each_list = []
                        each_list.append(JDateLOG)
                        each_list.append(record.Time_Log)
                        each_list.append(Parameter)
                        # append each query record as a list to lstRecords
                        lstRecords.append(each_list)

                # generate file name to save on Client Side
                file_name_gen = f"Scoplant_ExcelLog-{selected_device}-{JDateLOG}-{randint(0, 100)}"

                # This if elif conditions checks the Selcted Parameter
                if selectParameter == "Lux":
                    bg_color = "f3cf31"

                elif selectParameter == "Humidity":
                    bg_color = "46d4ff"

                elif selectParameter == "Temperature":
                    bg_color = "e90042"

                elif selectParameter == "Soil_Moisture":
                    bg_color = "1895ca"

                elif selectParameter == "Soil_tempurature":
                    bg_color = "912525"

                elif selectParameter == "EC":
                    bg_color = "26d8bd"

                elif selectParameter == "Total":
                    bg_colors = {
                        "Lux": "f3cf31",
                        "Humidity": "46d4ff",
                        "Temperature": "e90042",
                        "Soil_Moisture": "1895ca",
                        "Soil_Tempurature": "912525",
                        "EC": "26d8bd"
                    }
                if selectParameter == "Total":
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    Date_cell = sheet.cell(row=1, column=1)
                    Date_cell.value = "Date"

                    Time_cell = sheet.cell(row=1, column=2)
                    Time_cell.value = "Time"

                    Lux_cell = sheet.cell(row=1, column=3)
                    Lux_cell.value = "Lux"

                    Humidity_cell = sheet.cell(row=1, column=4)
                    Humidity_cell.value = "Humidity"

                    Temperature_cell = sheet.cell(row=1, column=5)
                    Temperature_cell.value = "Temperature"

                    Soil_Moisture_cell = sheet.cell(row=1, column=6)
                    Soil_Moisture_cell.value = "Soil Moisture"

                    Soil_Tempurature_cell = sheet.cell(row=1, column=7)
                    Soil_Tempurature_cell.value = "Soil Tempurature"

                    EC_cell = sheet.cell(row=1, column=8)
                    EC_cell.value = "EC"

                    sheet['A1'].fill = PatternFill(
                        start_color="d4c396", end_color="d4c396", fill_type="solid")
                    sheet['B1'].fill = PatternFill(
                        start_color="a6bdac", end_color="a6bdac", fill_type="solid")
                    sheet['C1'].fill = PatternFill(
                        start_color=bg_colors["Lux"], end_color=bg_colors["Lux"], fill_type="solid")
                    sheet['D1'].fill = PatternFill(
                        start_color=bg_colors["Humidity"], end_color=bg_colors["Humidity"], fill_type="solid")
                    sheet['E1'].fill = PatternFill(
                        start_color=bg_colors["Temperature"], end_color=bg_colors["Temperature"], fill_type="solid")
                    sheet['F1'].fill = PatternFill(
                        start_color=bg_colors["Soil_Moisture"], end_color=bg_colors["Soil_Moisture"], fill_type="solid")
                    sheet['G1'].fill = PatternFill(
                        start_color=bg_colors["Soil_Tempurature"], end_color=bg_colors["Soil_Tempurature"], fill_type="solid")
                    sheet['H1'].fill = PatternFill(
                        start_color=bg_colors["EC"], end_color=bg_colors["EC"], fill_type="solid")
                    Date_cell.font = Font(size=15, bold=True, color="000000")
                    Time_cell.font = Font(size=15, bold=True, color="000000")
                    Lux_cell.font = Font(size=15, bold=True, color="000000")
                    Humidity_cell.font = Font(
                        size=15, bold=True, color="000000")
                    Temperature_cell.font = Font(
                        size=15, bold=True, color="000000")
                    Soil_Moisture_cell.font = Font(
                        size=15, bold=True, color="000000")
                    Soil_Tempurature_cell.font = Font(
                        size=15, bold=True, color="000000")
                    EC_cell.font = Font(size=15, bold=True, color="000000")

                    num_row = 2
                    for row in lstRecords:
                        date_row = sheet.cell(row=num_row, column=1)
                        date_row.value = row[0]

                        time_row = sheet.cell(row=num_row, column=2)
                        time_row.value = row[1]

                        lux_row = sheet.cell(row=num_row, column=3)
                        lux_row.value = row[2]

                        humidity_row = sheet.cell(row=num_row, column=4)
                        humidity_row.value = row[3]

                        temperature_row = sheet.cell(row=num_row, column=5)
                        temperature_row.value = row[4]

                        soil_moisture_row = sheet.cell(row=num_row, column=6)
                        soil_moisture_row.value = row[5]

                        soil_tempurature_row = sheet.cell(
                            row=num_row, column=7)
                        soil_tempurature_row.value = row[6]

                        ec_row = sheet.cell(row=num_row, column=8)
                        ec_row.value = row[7]

                        num_row = num_row + 1

                    wb.save(f"temp/{file_name_gen}.xlsx")

                else:
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    title1 = sheet.cell(row=1, column=1)
                    title1.value = "Date"
                    title2 = sheet.cell(row=1, column=2)
                    title2.value = "Time"
                    title3 = sheet.cell(row=1, column=3)
                    title3.value = selectParameter
                    sheet['A1'].fill = PatternFill(
                        start_color=bg_color, end_color=bg_color, fill_type="solid")
                    sheet['B1'].fill = PatternFill(
                        start_color=bg_color, end_color=bg_color, fill_type="solid")
                    sheet['C1'].fill = PatternFill(
                        start_color=bg_color, end_color=bg_color, fill_type="solid")
                    title1.font = Font(size=15, bold=True, color="000000")
                    title2.font = Font(size=15, bold=True, color="000000")
                    title3.font = Font(size=15, bold=True, color="000000")

                    num_row = 2
                    for row in lstRecords:
                        date_row = sheet.cell(row=num_row, column=1)
                        date_row.value = row[0]
                        time_row = sheet.cell(row=num_row, column=2)
                        time_row.value = row[1]
                        value_row = sheet.cell(row=num_row, column=3)
                        value_row.value = row[2]
                        num_row = num_row + 1

                    wb.save(f"temp/{file_name_gen}.xlsx")

                # Sending file on Clint side to download
                with open(rf'temp\{file_name_gen}.xlsx', 'rb') as f:
                    try:
                        response = HttpResponse(f)
                        response['content_type'] = "application/octet-stream"
                        response['Content-Disposition'] = 'attachment; filename=' + \
                            os.path.basename(rf'temp\{file_name_gen}.xlsx')
                        return response
                    except Exception:
                        raise Http404

            # Checks if selectExport is PDFTableExport
            elif selectExport == "PDFTableExport":
                selectExport = "pdf- tabular"
                paramtererfor = ''
                if selectParameter == "Lux":
                    paramtererfor = 'light intensity'

                elif selectParameter == "Humidity":
                    paramtererfor = 'moisture'

                elif selectParameter == "Temperature":
                    paramtererfor = 'temperature'

                elif selectParameter == "Soil_Moisture":
                    paramtererfor = 'soil moisture'

                elif selectParameter == "Soil_tempurature":
                    paramtererfor = 'soil temperature'

                elif selectParameter == "EC":
                    paramtererfor = 'EC'

                elif selectParameter == "Total":
                    paramtererfor = 'output all data'

                context = {
                    'title': 'output PDF',
                    'startDateJalali': request.POST.get("start_date"),
                    'endDateJalali': request.POST.get("end_date"),
                    'export_method': selectExport,
                    'export_parameter': selectParameter,
                    'requested_by': request.user,
                    'requested_for': selected_device,
                    'qs_res': LogsQS,
                    'selected_parameter': paramtererfor
                }

                # Set the PDF Template for export
                template = get_template('pdf_table.html')
                # Set the info to pass on template
                html = template.render(context)
                # generate file name to save on Client Side
                file_name_gen = f"Scoplant_PDFLog-{selected_device}-{selectParameter}-{randint(0, 100)}"

                # Create file on server
                file = open(f'temp/{file_name_gen}.pdf', "w+b")
                # Set Encoding and other values
                # pisa.CreatePDF(html.encode('utf-8'),
                #                dest=file, encoding='utf-8')
                file.seek(0)
                # set file info on pdf var
                pdf = file.read()
                # Close the file
                file.close()
                # Remove the file from server
                os.remove(file.name)

                # Sending file on Clint side to download
                response = HttpResponse(pdf)
                response['content_type'] = "application/pdf"
                response['Content-Disposition'] = 'attachment; filename=' + \
                    os.path.basename(rf'temp\{file_name_gen}.pdf')
                return response

            # Checks if selectExport is PDFChartExportParts
            elif selectExport == "PDFChartExportParts":
                selectExport = "جدول جزئی -pdf"
                lstRecords = []
                for record in LogsQS:
                    # Convert gregorian date into jalali date
                    DateLOG = str(record.Date_Log)
                    JDateLOG = Jalaali.to_jalaali(
                        int(DateLOG[0:4]), int(DateLOG[5:7]), int(DateLOG[8:10]))
                    JDateLOG = f"{JDateLOG['jy']}-{JDateLOG['jm']}-{JDateLOG['jd']}"
                    each_data = f"{record.Date_Log}, {record.Time_Log}, {record.Lux_Log}, {record.Humidity_Log}, {record.Temperature_Log}, {record.SoilMoisture_Log}, {record.SoilTemperature_Log}, {record.EC_Log}"

                    lstRecords.append(each_data)

                # generate file name to save on Client Side
                file_name_gen = f"Scoplant_ChartPDF-{selected_device}-{JDateLOG}-{randint(0, 100)}"

                with open(f'temp/{file_name_gen}.csv', 'w') as f:
                    f.write(
                        "Date,Time,Lux,Humidity,Temperature,Soil_Moisture,Soil_tempurature,EC")
                    f.write("\n")
                    for each in lstRecords:
                        f.write(each)
                        f.write("\n")

                df = pd.read_csv(f'temp/{file_name_gen}.csv')

                if selectParameter == "Lux":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='light intensity')
                    fig.update_traces(line_color='#f3cf31')
                    # fig.update_layout(
                    #     font=dict(
                    #         size=18,
                    #         color="Black"
                    #     )
                    # )

                elif selectParameter == "Humidity":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='Humidity')
                    fig.update_traces(line_color='#46d4ff')

                elif selectParameter == "Temperature":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='temperature')
                    fig.update_traces(line_color='#e90042')

                elif selectParameter == "Soil_Moisture":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='soil moisture')
                    fig.update_traces(line_color='#1895ca')

                elif selectParameter == "Soil_tempurature":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='soil temperature')
                    fig.update_traces(line_color='#912525')

                elif selectParameter == "EC":
                    fig = px.line(df, x='Date', y=f'{selectParameter}', range_x=[
                                  gregorianStartDate, gregorianEndtDate])
                    fig.update_yaxes(title_text='EC')
                    fig.update_traces(line_color='#26d8bd')

                elif selectParameter == "Total":
                    total_data = pd.pivot_table(df, values=[
                                                'Lux', 'Humidity', 'Temperature', 'Soil_Moisture', 'Soil_tempurature', 'EC'], index='Date')
                    trace_lux = go.Scatter(
                        x=total_data.index,
                        y=total_data.Lux,
                        mode='lines',
                        name='Lux'
                    )
                    trace_humidity = go.Scatter(
                        x=total_data.index,
                        y=total_data.Humidity,
                        mode='lines',
                        name='Humidity'
                    )
                    trace_temperature = go.Scatter(
                        x=total_data.index,
                        y=total_data.Temperature,
                        mode='lines',
                        name='Temperature'
                    )
                    trace_soil_moisture = go.Scatter(
                        x=total_data.index,
                        y=total_data.Soil_Moisture,
                        mode='lines',
                        name='Soil_Moisture'
                    )
                    trace_soil_tempurature = go.Scatter(
                        x=total_data.index,
                        y=total_data.Soil_tempurature,
                        mode='lines',
                        name='Soil_tempurature'
                    )
                    trace_ec = go.Scatter(
                        x=total_data.index,
                        y=total_data.EC,
                        mode='lines',
                        name='EC'
                    )

                    all_chart_data = [
                        trace_lux, trace_humidity, trace_temperature, trace_soil_moisture, trace_soil_tempurature, trace_ec]

                    layout = go.Layout(title='Total Chart Export')

                    fig = go.Figure(data=all_chart_data, layout=layout)

                os.remove(f'temp/{file_name_gen}.csv')
                plotly.io.write_image(
                    fig, f'temp/{file_name_gen}.pdf', format='pdf')

                file_server = pathlib.Path(f'temp/{file_name_gen}.pdf')
                if not file_server.exists():
                    raise Http404()
                else:
                    file_to_download = open(str(file_server), 'rb')
                    response = FileResponse(
                        file_to_download, content_type='application/force-download')
                    response['Content-Disposition'] = f'inline; filename="{file_name_gen}.pdf"'
                    return response

        # else data range is not valid or wrong
        else:
            wrong_date_range = True

    number_of_devices = 0
    for item in device_info:
        if item.User == user:
            number_of_devices = number_of_devices+1

    avatar_gen = f"/site_statics/images/Avatars/Avatar {randint(1,12)}.jpg"
    context = {
        'device_info': device_info,
        'number_of_devices': number_of_devices,
        'user': user,
        'title': "Device Reporting",
        'export_methods': export_methods,
        'wrong_date_range': wrong_date_range,
        'avatar_gen_url': avatar_gen,
    }
    return render(request, 'reporting.html', context)
